"""One-shot: copy GLBs + generate src/data/trophies.ts from Excel research."""
from __future__ import annotations

import re
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(r'D:\AhmedRashad\FirstOption\FirstOption\ALAHLY')
SRC_DIR = ROOT / 'trophies' / '3D-models'
DST_DIR = ROOT / 'football-club-webar' / 'public' / 'models' / 'trophies'
XLSX = ROOT / 'Al_Ahly_Trophies_Research_FIXED.xlsx'
OUT_TS = ROOT / 'football-club-webar' / 'src' / 'data' / 'trophies.ts'

# Canonical competition keys (match Excel as closely as possible)
CANONICAL = [
    'Egyptian Premier League',
    'Egypt Cup',
    'Egyptian Super Cup',
    'Sultan Hussein Cup',
    'Cairo League',
    'CAF Champions League',
    'CAF Confederation Cup',
    'CAF Super Cup',
    'FIFA African–Asian–Pacific Cup',
    'Afro-Asian Club Championship',
    'Arab Club Champions Cup',
    "Arab Cup Winners' Cup",
    'FIFA Club World Cup',
]

SLUG = {
    'Egyptian Premier League': 'egyptian-premier-league',
    'Egypt Cup': 'egypt-cup',
    'Egyptian Super Cup': 'egyptian-super-cup',
    'Sultan Hussein Cup': 'sultan-hussein-cup',
    'Cairo League': 'cairo-league',
    'CAF Champions League': 'caf-champions-league',
    'CAF Confederation Cup': 'caf-confederation-cup',
    'CAF Super Cup': 'caf-super-cup',
    'FIFA African–Asian–Pacific Cup': 'fifa-african-asian-pacific-cup',
    'Afro-Asian Club Championship': 'afro-asian-club-championship',
    'Arab Club Champions Cup': 'arab-club-champions-cup',
    "Arab Cup Winners' Cup": 'arab-cup-winners-cup',
    'FIFA Club World Cup': 'fifa-club-world-cup',
}

NAME_AR = {
    'Egyptian Premier League': 'الدوري المصري الممتاز',
    'Egypt Cup': 'كأس مصر',
    'Egyptian Super Cup': 'كأس السوبر المصري',
    'Sultan Hussein Cup': 'كأس السلطان حسين',
    'Cairo League': 'دوري القاهرة',
    'CAF Champions League': 'دوري أبطال أفريقيا',
    'CAF Confederation Cup': 'كأس الكونفيدرالية الأفريقية',
    'CAF Super Cup': 'كأس السوبر الأفريقي',
    'FIFA African–Asian–Pacific Cup': 'كأس الفيفا الأفريقية الآسيوية الباسيفيكية',
    'Afro-Asian Club Championship': 'بطولة الأندية الأفروآسيوية',
    'Arab Club Champions Cup': 'كأس العرب للأندية الأبطال',
    "Arab Cup Winners' Cup": 'كأس الكؤوس العربية',
    'FIFA Club World Cup': 'كأس العالم للأندية',
}

SUMMARIES = {
    'Egyptian Premier League': 'العمود الفقري لهيمنة الأهلي المحلية، بأكبر رصيد ألقاب في تاريخ الدوري المصري.',
    'Egypt Cup': 'مجد الكأس المحلية عبر عقود من الانتصارات في الأدوار الإقصائية.',
    'Egyptian Super Cup': 'مواجهة الأبطال محلياً بين بطل الدوري وبطل الكأس.',
    'Sultan Hussein Cup': 'بطولة تاريخية من عصر ما قبل الدوري الحديث، شهدت بدايات مجد الأهلي.',
    'Cairo League': 'بطولة القاهرة التاريخية التي شكّلت جزءاً مهماً من سجل الأهلي قبل التوحيد الوطني.',
    'CAF Champions League': 'أعظم بطولات القارة الأفريقية، ورمز سيادة الأهلي القارية.',
    'CAF Confederation Cup': 'لقب قاري يؤكد عمق مشاركة الأهلي في منافسات الكاف.',
    'CAF Super Cup': 'مواجهة أبطال أفريقيا، وتاج موسمي يضاف إلى خزائن الأهلي.',
    'FIFA African–Asian–Pacific Cup': 'لقب فيفا نادر يربط أبطال أفريقيا وآسيا والباسيفيك.',
    'Afro-Asian Club Championship': 'بطولة تاريخية جمعت أبطال أفريقيا وآسيا في مواجهة عابرة للقارات.',
    'Arab Club Champions Cup': 'تتويج عربي يؤكد مكانة الأهلي بين أندية المنطقة.',
    "Arab Cup Winners' Cup": 'لقب عربي تاريخي من بطولة أبطال الكؤوس.',
    'FIFA Club World Cup': 'المنافسة العالمية للأندية؛ سجل الأهلي يشمل مراكز برونزية موثّقة.',
}

STATUS_AR = {
    'Active': 'نشطة',
    'Defunct': 'منتهية',
    'Active format': 'صيغة نشطة',
    'Historical format': 'صيغة تاريخية',
    'Competition format changed': 'تغيّرت صيغة البطولة',
}


def norm(s: str) -> str:
    s = s.replace('\u2013', '-').replace('\u2014', '-').replace('\u2011', '-')
    s = s.replace("'", "'").replace("'", "'")
    return re.sub(r'\s+', ' ', s.strip()).lower()


def resolve_canonical(name: str) -> str | None:
    n = norm(name)
    for c in CANONICAL:
        if norm(c) == n:
            return c
    if 'african' in n and 'pacific' in n:
        return 'FIFA African–Asian–Pacific Cup'
    if 'arab' in n and 'winners' in n:
        return "Arab Cup Winners' Cup"
    if 'club world' in n:
        return 'FIFA Club World Cup'
    if 'afro-asian' in n or 'afroasian' in n.replace('-', ''):
        return 'Afro-Asian Club Championship'
    return None


def category_ar(raw: str) -> str:
    n = norm(raw)
    mapping = {
        'domestic - major': 'محلي — بطولات كبرى',
        'domestic - historical': 'محلي — تاريخي',
        'africa - caf': 'أفريقيا — الكاف',
        'intercontinental / fifa': 'قاري / فيفا',
        'intercontinental - historical': 'قاري — تاريخي',
        'arab / regional': 'عربي / إقليمي',
        'fifa / world - placement': 'فيفا / العالم — مراكز',
    }
    return mapping.get(n, raw)


def achievement_ar(text: str, rank: int) -> str:
    t = (text or '').lower()
    if rank == 3 or 'third' in t or 'bronze' in t:
        return 'المركز الثالث'
    if 'champion' in t or rank == 1:
        return 'بطل'
    return text or 'بطل'


def slugify_filename(name: str) -> str:
    s = name.lower()
    s = s.replace("'", '').replace("'", '').replace("'", '')
    s = s.replace('_', '-').replace('\u2013', '-').replace('\u2014', '-')
    s = re.sub(r'[^a-z0-9]+', '-', s)
    return s.strip('-') + '.glb'


def copy_glbs() -> dict[str, str]:
    """Return map canonical competition -> public modelSrc."""
    DST_DIR.mkdir(parents=True, exist_ok=True)
    # Build lookup from source files
    src_files = list(SRC_DIR.glob('*.glb'))
    print('Source GLBs:')
    for p in src_files:
        print(' ', repr(p.name))

    model_src: dict[str, str] = {}
    for canon in CANONICAL:
        slug = SLUG[canon]
        dest = DST_DIR / f'{slug}.glb'
        # Find matching source
        target_norm = norm(canon)
        match = None
        for p in src_files:
            stem = p.stem
            if resolve_canonical(stem) == canon or norm(stem) == target_norm:
                match = p
                break
            # FIFA African special
            if canon.startswith('FIFA African') and 'african' in norm(stem) and 'pacific' in norm(stem):
                match = p
                break
            if canon == 'FIFA Club World Cup' and 'club-world' in norm(stem).replace(' ', '-'):
                match = p
                break
            if canon == 'Arab Club Champions Cup' and 'arab' in norm(stem) and 'champions' in norm(stem):
                match = p
                break
            if canon == "Arab Cup Winners' Cup" and 'arab' in norm(stem) and 'winners' in norm(stem):
                match = p
                break
        if not match:
            raise SystemExit(f'No GLB found for {canon!r}')
        shutil.copy2(match, dest)
        model_src[canon] = f'/models/trophies/{slug}.glb'
        print(f'COPIED {match.name!r} -> {dest.name}')

    return model_src


def ts_escape(s: str) -> str:
    return s.replace('\\', '\\\\').replace("'", "\\'")


def generate_ts(model_src: dict[str, str]) -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # Competitions sheet
    ws = wb['Competitions']
    rows = list(ws.iter_rows(values_only=True))
    hdr = next(i for i, r in enumerate(rows) if r and r[0] == 'Category')
    comps: dict[str, dict] = {}
    for r in rows[hdr + 1 :]:
        if not r[1]:
            continue
        canon = resolve_canonical(str(r[1]))
        if not canon:
            continue
        comps[canon] = {
            'categoryAr': category_ar(str(r[0] or '')),
            'officialTitles': int(r[2] or 0),
            'statusAr': STATUS_AR.get(str(r[3] or '').strip(), str(r[3] or '').strip()),
        }

    # Achievements sheet
    ws2 = wb['Achievements']
    rows2 = list(ws2.iter_rows(values_only=True))
    hdr2 = next(i for i, r in enumerate(rows2) if r and r[0] == 'Category')
    achs: dict[str, list] = {c: [] for c in CANONICAL}
    for r in rows2[hdr2 + 1 :]:
        if not r[1]:
            continue
        canon = resolve_canonical(str(r[1]))
        if not canon or canon not in achs:
            continue
        season = str(r[2]).strip() if r[2] is not None else ''
        exact = str(r[3]).strip() if r[3] is not None else ''
        rank = int(r[4]) if r[4] is not None else 1
        ach_text = str(r[5]).strip() if r[5] is not None else ''
        tnum = int(r[6]) if r[6] is not None else None
        entry = {
            'season': season,
            'exactDate': exact or None,
            'rank': rank,
            'achievementAr': achievement_ar(ach_text, rank),
            'trophyNumber': tnum,
        }
        achs[canon].append(entry)

    lines: list[str] = []
    lines.append('export interface TrophyAchievement {')
    lines.append('  season: string')
    lines.append('  exactDate?: string')
    lines.append('  rank: number')
    lines.append('  achievementAr: string')
    lines.append('  trophyNumber: number | null')
    lines.append('}')
    lines.append('')
    lines.append('export interface TrophyDefinition {')
    lines.append('  id: string')
    lines.append('  nameEn: string')
    lines.append('  nameAr: string')
    lines.append('  categoryAr: string')
    lines.append('  statusAr: string')
    lines.append('  officialTitles: number')
    lines.append('  summaryAr: string')
    lines.append('  modelSrc: string')
    lines.append('  achievements: TrophyAchievement[]')
    lines.append('}')
    lines.append('')
    lines.append('/**')
    lines.append(' * Trophy cabinet — competitions with 3D models.')
    lines.append(' * Dialog copy is Arabic; ids/nameEn are for code and ExplorePanel.')
    lines.append(' */')
    lines.append('export const trophies: TrophyDefinition[] = [')

    for canon in CANONICAL:
        meta = comps.get(canon)
        if not meta:
            raise SystemExit(f'Missing competition row for {canon!r}')
        tid = SLUG[canon]
        lines.append('  {')
        lines.append(f"    id: '{tid}',")
        lines.append(f"    nameEn: '{ts_escape(canon)}',")
        lines.append(f"    nameAr: '{ts_escape(NAME_AR[canon])}',")
        lines.append(f"    categoryAr: '{ts_escape(meta['categoryAr'])}',")
        lines.append(f"    statusAr: '{ts_escape(meta['statusAr'])}',")
        lines.append(f"    officialTitles: {meta['officialTitles']},")
        lines.append(f"    summaryAr: '{ts_escape(SUMMARIES[canon])}',")
        lines.append(f"    modelSrc: '{model_src[canon]}',")
        lines.append('    achievements: [')
        for a in achs[canon]:
            lines.append('      {')
            lines.append(f"        season: '{ts_escape(a['season'])}',")
            if a['exactDate']:
                lines.append(f"        exactDate: '{ts_escape(a['exactDate'])}',")
            lines.append(f"        rank: {a['rank']},")
            lines.append(f"        achievementAr: '{ts_escape(a['achievementAr'])}',")
            tnum = a['trophyNumber']
            lines.append(f"        trophyNumber: {tnum if tnum is not None else 'null'},")
            lines.append('      },')
        lines.append('    ],')
        lines.append('  },')

    lines.append(']')
    lines.append('')
    lines.append('export function getTrophyById(id: string): TrophyDefinition | undefined {')
    lines.append('  return trophies.find((t) => t.id === id)')
    lines.append('}')
    lines.append('')
    lines.append('export function getTrophyIndex(id: string): number {')
    lines.append("  return trophies.findIndex((t) => t.id === id)")
    lines.append('}')
    lines.append('')

    OUT_TS.write_text('\n'.join(lines), encoding='utf-8')
    print(f'Wrote {OUT_TS}')
    for canon in CANONICAL:
        print(f'  {SLUG[canon]}: {len(achs[canon])} achievements, titles={comps[canon]["officialTitles"]}')


if __name__ == '__main__':
    models = copy_glbs()
    generate_ts(models)
